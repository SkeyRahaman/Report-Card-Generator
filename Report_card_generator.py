from datetime import datetime, date, timedelta
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import pandas as pd
import numpy as np
import openpyxl
import requests
import seaborn as sns
import matplotlib.pyplot as plt
import os
import difflib
import smtplib
from photos_and_other_requirement import email_credentials
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from string import Template


class Main:
    def __init__(self):
        self.cwd = os.getcwd()
        self.Make_Directorys(self.cwd + "\\" + "Data")
        self.Clone_the_dataset_to_this_machine("data\\" + 'Student Gradebook.xlsx', "https://docs.google.com/spreadsheets/d/1HYjfEe3aCbufbqIXKs0Xz-gfoQNztGhCN1ivx0gZXnc/export?format = xlsx")
        self.Clone_the_dataset_to_this_machine("data\\" + 'users.csv', "https://docs.google.com/spreadsheets/d/1hIj8fA3VeozglJcDXkN_icSbhoyhMiycc9cSp_PYMUY/export?format=csv")
        data = self.creat_data("Data\\" + "Student Gradebook.xlsx")
        self.user_data = self.Get_user_data()
        data = self.Data_cleaning(data)
        self.data = self.Add_Month_column(data)
        self.data = self.Add_heighest_marks_column(self.data)
        self.Make_Directorys(self.cwd + "\\" + "Report_card")

        given_month = (date.today().replace(day=1) - timedelta(days=1)).strftime("%B")
        for i in list(self.user_data["Df_name"].value_counts().index):
            given_name = i
            self.Make_Directorys(self.cwd + "\\" + "Report_card\\" + given_month)
            self.Make_Directorys(
                self.cwd + "\\" + "Report_card\\" + given_month + "\\" + given_name + "_" + given_month)
            self.file_loc = self.cwd + "\\" + "Report_card\\" + given_month + "\\" + given_name + "_" + given_month + "\\"
            self.Start_making_pdf_of(given_name, given_month)

        #self.Send_mail_login(given_month)

    def Make_Directorys(self, path):
        try:
            os.mkdir(path)
        except:
            print("Can't create the folder with path", "  ", path)

    def Get_user_data(self):
        return pd.read_csv(self.cwd + "\\Data\\" + "users.csv")

    def Clone_the_dataset_to_this_machine(self, path, url):
        a = requests.get(url)
        resp = requests.get(url)
        output = open(path, 'wb')
        output.write(resp.content)
        output.close()

    def creat_data(self, file):
        wb = openpyxl.load_workbook(file)
        Number_of_sheets = len(wb.sheetnames)
        for i in range(len(wb.sheetnames)):
            try:
                data = pd.concat([data, pd.read_excel(file, sheet_name=i)])
            except:
                data = pd.read_excel(file, sheet_name=0)
        return data

    def Add_Month_column(self, data):
        date_list = []
        day_list = []
        year_list = []
        for i in data["Date"]:
            try:
                date_list.append(i.strftime("%B"))
                day_list.append(i.strftime("%d"))
                year_list.append(i.strftime("%Y"))
            except:
                try:
                    date_list.append(datetime.strptime(str(i), "%d/%m/%Y").strftime("%B"))
                    day_list.append(datetime.strptime(str(i), "%d/%m/%Y").strftime("%d"))
                    year_list.append(datetime.strptime(str(i), "%d/%m/%Y").strftime("%Y"))
                except:
                    print("Interrupt in a line number : - ", a, "which is having a entry of:-", i)
                    break
        data["Date(only)"] = day_list
        data["Month"] = date_list
        data["Year"] = year_list
        return data

    #CALCULATION OF HIGHEST MARKS

    def Add_heighest_marks_column(self, data):
        a = data['Task'].value_counts().index
        highest = []
        for i in a:
            highest.append(max(data[data['Task'] == i]['Points']))
        data_highest_table = {'Task': a, 'Highest': highest}
        data_high = pd.DataFrame(data_highest_table)
        return data.merge(data_high, on='Task')

    #DATA CLEANING

    def Data_cleaning(self, data):
        name_and_def_name = {}
        corrected_name = []
        for i in range(len(self.user_data)):
            name_and_def_name[self.user_data['Full_name'][i]] = self.user_data['Df_name'][i]

        all_full_names = list(name_and_def_name.keys())
        names = list(self.user_data["Df_name"])
        for i in data["Student"].values:
            if i in names:
                corrected_name.append(i)
            else:
                suggested_name = difflib.get_close_matches(str(i), all_full_names, cutoff=0.5)
                if len(suggested_name) == 0:
                    corrected_name.append("Not_a_student")
                else:
                    corrected_name.append(name_and_def_name[suggested_name[0]])
        data["Student"] = corrected_name
        return data[data["Student"] != "Not_a_student"]

    def Start_making_pdf_of(self, name, month):
        self.Full_Name = str(self.user_data[self.user_data["Df_name"] == name].iloc[0]["Full_name"])
        self.Name = name
        self.College = str(self.user_data[self.user_data["Df_name"] == name].iloc[0]["college"]).split(",")[0]
        self.Year = "First"
        self.Month = month
        self.Date = str(date.today().strftime("%d %B  %Y"))
        self.Email = str(self.user_data[self.user_data["Df_name"] == name].iloc[0]["email"])
        self.Number_of_task_wins = self.number_of_task_wins(self.Name, self.Month)
        self.Rank_among_the_class = self.rank_of_the_student(self.Name, self.Month)
        self.Late_submition_ratio = self.late_Submition_Ratio(self.Name, self.Month)
        self.Percentage = self.percentage_of_the_student(self.Name, self.Month)
        self.Percentile = self.percentile_of_the_student(self.Name, self.Month)
        self.Table_Content = self.table_Content(self.Name, self.Month)
        self.Table_summary = self.table_summary(self.Name, self.Month)
        self.main_of_pdf(self.Name, self.Month)

    # CALCULATION OF NUMBER OF TASK WINNER

    def number_of_task_wins(self, name, month):
        working_data = self.data
        working_data = working_data[working_data["Student"] == name]
        working_data = working_data[working_data["Month"] == month]
        working_data = working_data[working_data["Task Winner"] == 1]
        return str(working_data['id'].count())

     # CALCULATION OF RANK OF STUDENT

    def rank_of_the_student(self, name, month):
        data_working = self.data[self.data['Month'] == month]
        data_task = data_working[data_working['Module'] != 'Ritual']
        data_rank = data_task.groupby('Student')['Points'].sum().sort_values(ascending=False).reset_index()
        inx = data_rank['Student'].tolist()
        i = pd.Index(inx).get_loc(name)
        return str(i + 1)

    # CALCULATION OF LATE SUBMISSION RATIO

    def late_Submition_Ratio(self, name, month):
        data_working = self.data[self.data['Month'] == month]
        data_student = data_working[data_working['Student'] == name]
        p = len(data_working[data_working['Module'] != 'Ritual']['Task'].value_counts().index)
        q = len(data_student[data_student['Late Submission'] == 1])
        ratio = q / p
        return str(round(ratio, 2))

    # CALCULATION OF PERCENTAGE OF THE STUDENT

    def percentage_of_the_student(self, name, month):
        data_working = self.data[self.data['Month'] == month]
        data_student = data_working[data_working['Student'] == name]
        data_task = data_student[data_student['id'] != 0]
        a = data_task['Points'].sum()
        b = data_task['Total'].sum()
        return str(round((a / b) * 100, 1))

    # CALCULATION OF PERCENTILE OF THE STUDENT

    def percentile_of_the_student(self, name, month):
        a = int(self.rank_of_the_student(name, month))
        b = len(self.data['Student'].value_counts().index)
        return str(round(((b - a) / b) * 100, 1))

    # DESIGNING A TABLE WITH RESULT ANALYSIS

    def table_Content(self, name, month):

        data_working = self.data[self.data['Month'] == month]
        data_working = data_working[data_working["Module"] != "Ritual"]
        data_student = data_working[data_working['Student'] == name]
        x1 = data_student.pivot_table(index='Module', values='Points', aggfunc='sum')
        y1 = data_student.pivot_table(index='Module', values='Total', aggfunc='sum')
        w1 = data_student.pivot_table(index='Module', values='Highest', aggfunc='sum')
        z1 = pd.concat([x1, y1, w1], axis=1)
        subject_percentile = [x1['Points'] / y1['Total'] * 100]
        z1['Percentile'] = " "
        for i in range(len(z1)):
            z1['Percentile'][i] = round(subject_percentile[0][i],1)
        z1.to_html('Data\\z_total.html')
        df = pd.read_html('Data\\z_total.html')
        return df[0].values.tolist()

    # DESIGNING OF TABLE SUMMARY

    def table_summary(self, name, month):
        data_working = self.data[self.data['Month'] == month]
        data_working = data_working[data_working["Module"] != "Ritual"]
        data_student = data_working[data_working['Student'] == name]
        x1 = data_student.pivot_table(index='Module', values='Points', aggfunc='sum')
        y1 = data_student.pivot_table(index='Module', values='Total', aggfunc='sum')
        w1 = data_student.pivot_table(index='Module', values='Highest', aggfunc='sum')
        z1 = pd.concat([x1, y1, w1], axis=1)
        subject_percentile = [x1['Points'] / y1['Total'] * 100]
        z1['Percentile'] = " "
        for i in range(len(z1)):
            z1['Percentile'][i] = subject_percentile[0][i]
        j = z1['Points'].sum()
        k = None
        l = z1['Highest'].sum()
        return list(["Total", j, k, l, self.percentage_of_the_student(name, month)])

    def main_of_pdf(self, name, month):
        c = canvas.Canvas((self.file_loc + name + "_" + month + ".pdf"), bottomup=1, pagesize=A4)
        c = self.draw_border(c, 35)
        c = self.draw_border(c, 32.5)
        c = self.draw_border(c, 30)
        c = self.draw_intro(c, 25, name)
        c = self.draw_table(c)
        c = self.draw_comparison_table(c)
        c = self.draw_acknowledgement(c, 22)
        c.showPage()
        c.save()

    def draw_border(self, c, m):
        c.line(m, m, 595.27 - m, m)
        c.line(m, 841.89 - m, 595.27 - m, 841.89 - m)
        c.line(m, m, m, 841.89 - m)
        c.line(595.27 - m, m, 595.27 - m, 841.89 - m)
        return c

    def draw_intro(self, c, Spacing, name):
        c.setFont('Times-Bold', 28)
        c.setFillColorRGB(0, 0, 0.77)
        c.drawCentredString(595.27 / 2 + 50, 750, text='CampusX Mentorship Programme')
        c.setFillColorRGB(0, 0, 0)
        c.setFont('Times-Roman', 22)
        c.drawCentredString(320, 720, text='Machine Learning')
        c.setFont('Times-Roman', 18)
        c.drawString(45, 680, ('NAME:-' + self.Full_Name))
        c.drawString(45, 680 - Spacing, 'COLLEGE:-' + self.College)
        # c.drawString(285, 680-2*Spacing, 'YEAR:-'+info.Year)
        c.drawString(45, 680 - 2 * Spacing, 'MONTH:-' + self.Month)
        c.drawString(45, 680 - 3 * Spacing, 'Email Address:- ' + self.Email)
        c.line(35, 680 - 3.5 * Spacing, 560.27, 680 - 3.5 * Spacing)
        c.drawInlineImage(image=(self.cwd + "\\photos_and_other_requirement\\" + "campusX_Final.jpg"), x=45, y=700, width=80, height=100)
        c.drawInlineImage(image=(self.cwd + "\\photos_and_other_requirement\\Student_photo\\" + name + ".jpg"), x=425, y=600, width=115,
                          height=115)
        return c

    def draw_table(self, c):
        c.drawInlineImage(image=(self.cwd + "\\photos_and_other_requirement\\" + "TABLE_MODULES.jpg"), x=45, y=410, width=500, height=180)
        c.setFont('Times-Bold', 10)
        Heading = ['MODULE', "Your Marks", 'Full Marks', 'Highest Marks', 'Percentage']
        for i in range(len(Heading)):
            c.drawCentredString(95 + i * 100, 575, Heading[i])
        writing_row = 555
        data = self.Table_Content
        for i in range(len(self.Table_Content)):
            for j in range(len(self.Table_Content[i])):
                c.drawCentredString(95 + j * 100, writing_row, str(data[i][j]))
            writing_row -= 20
        for i in range(len(self.Table_summary)):
            if str(self.Table_summary[i]) == "None":
                continue
            c.drawCentredString(95 + i * 100, 415, str(self.Table_summary[i]))
        return c

    def draw_comparison_table(self, c):
        Y = 160  # Y scale of the second and third graph
        self.Give_me_first_graph_for_the_month_of(self.Name, self.Month)
        self.Creat_spided_plot(self.Name, self.Month)
        c.drawInlineImage(image=(self.file_loc + self.Name + "_" + self.Month + "_" + "graph1.jpg"), x=45, y=Y,
                          width=275, height=180)
        c.drawInlineImage(image=(self.file_loc + self.Name + "_" + self.Month + "Graph2.jpg"), x=335, y=Y, width=200,
                          height=180)
        c.setFont('Times-Roman', 19)
        c.drawString(45, 390, 'Comparison between this month and average till now on the ')
        c.drawString(55, 370, 'basic of:')
        c.drawString(65, 350, '1. Task Subject:')
        c.drawString(355, 350, '2. Task Value:')
        return c

    def Give_me_first_graph_for_the_month_of(self, Name, Month_name):
        working_data = self.data
        working_data = working_data[working_data["Student"] == Name]
        data_for_graph_one = working_data.groupby("Module")['Points', 'Total'].sum().reset_index()
        data_for_graph_one["Percentage"] = round(data_for_graph_one["Points"] / data_for_graph_one['Total'] * 100)
        data_for_graph_one["For_the_month_of"] = "Till Now"
        data_for_given_month = working_data[working_data["Month"] == Month_name]
        data_for_given_month = data_for_given_month.groupby("Module")['Points', 'Total'].sum().reset_index()
        data_for_given_month["Percentage"] = round(data_for_given_month["Points"] / data_for_given_month['Total'] * 100)
        data_for_given_month["For_the_month_of"] = str(Month_name)
        out_put_dataframe = pd.concat([data_for_given_month, data_for_graph_one])
        out_put_dataframe = out_put_dataframe[out_put_dataframe["Module"] != "Ritual"]
        plot = sns.barplot(x='Percentage', y='Module', data=out_put_dataframe, hue='For_the_month_of',
                           palette=["#00ff00", "Green"]) \
            .legend(loc='lower left', bbox_to_anchor=(0.7, 1.0))
        plot.get_figure().savefig((self.file_loc + Name + "_" + Month_name + "_" + "graph1.jpg"), dpi=300,

                                  bbox_inches='tight')
        del(plot)

    def Creat_spided_plot(self, name, month):
        working_data = self.data
        student_data = working_data[working_data["Student"] == name]
        Current_month_data = student_data[student_data["Month"] == month]
        one = self.return_me_the_df(Current_month_data)
        two = self.return_me_the_df(student_data)
        three = two[two['Type'].isin(one['Type'].values)]
        one.index = one['Type']
        one.drop(columns={'Type'}, inplace=True)
        three.index = three['Type']
        three.drop(columns={'Type'}, inplace=True)
        three.rename(columns={'Points': 'Points_all', 'FM': 'FM_all'}, inplace=True)
        four = pd.concat([one, three], axis=1, )
        four['per'] = (four['Points'] / four['FM'] * 100).values
        four['per_all'] = round(four['Points_all'] / four['FM_all'] * 100).values
        four.drop(columns={'Points', 'FM', 'Points_all', 'FM_all'}, inplace=True)
        self.plot(four, name, month)

    def return_me_the_df(self, data):
        My_list = []
        for i in range(len(data["id"])):
            current_row = data.iloc[i].values
            jata = [current_row[4], current_row[7], current_row[8]]
            My_list.append(jata)
        transfer = My_list
        next_my_list = []
        for i in transfer:
            a = i
            w = a[0].replace(" ", "").split(",")
            for j in w:
                one = j
                two = a[1]
                three = a[2]
                four = [one, two, three]
                next_my_list.append(four)

        spider_plot_df = pd.DataFrame(next_my_list)
        spider_plot_df.rename(columns={0: "Type", 1: "Points", 2: "FM"}, inplace=True)
        main_result = spider_plot_df.groupby("Type")["Points", "FM"].sum().reset_index()
        return main_result

    def plot(self, data, name, month):
        labels = np.array(data.index)
        stats = data.values
        angles = np.linspace(0, 2 * np.pi, len(labels), endpoint=False)
        stats = np.concatenate((stats, [stats[0]]))
        angles = np.concatenate((angles, [angles[0]]))
        ax = plt.subplot(111, polar=True)
        plt.xticks(angles[:-1], labels)
        ax.plot(angles, stats[:, 1], "o-", linewidth=3, color='Green', label="Till Now")
        ax.plot(angles, stats[:, 0], "o-", linewidth=1, color='#00ff00', label=str(month))
        ax.fill(angles, stats, 'teal', alpha=0.1)
        ax.set_title("")
        plt.legend(loc='lower right', bbox_to_anchor=(0.1, 1.0))
        ax.get_figure().savefig((self.file_loc + name + "_" + month + "Graph2.jpg"), dpi= 200, bbox_inches='tight')
        plt.close()

    def draw_acknowledgement(self, c, Spacing):
        c.line(35, 45 + 5 * Spacing, 560.27, 45 + 5 * Spacing)

        c.setFont('Times-Roman', 18)
        c.drawString(45, 45 + 4 * Spacing, 'Number of task wins:-' + self.Number_of_task_wins)
        c.drawString(45, 45 + 3 * Spacing, 'Rank among the class:-' + self.Rank_among_the_class)
        c.drawString(45, 45 + 2 * Spacing, 'Late submission ratio:-' + self.Late_submition_ratio)


        c.drawString(300, 45 + 4 * Spacing, 'Overall percentage :-' + self.Percentage + "%")
        c.drawString(300, 45 + 3 * Spacing, 'Overall percentile:-' + self.Percentile)
        c.drawString(300, 45 + 2 * Spacing, 'Generated on:-' + self.Date)

        c.setFont('Times-Roman', 10)
        c.drawString(45, 45 + Spacing, '*This is a computer-generated document. No signature is required.')
        c.drawString(45, 45, 'For any queries:-  https://mywbut.com')


        return c

    def Send_mail_login(self, month):
        server = smtplib.SMTP("smtp.gmail.com:587")
        login= 0
        server.ehlo()
        server.starttls()
        server.ehlo()
        try:
            server.login(email_credentials.EMAIL_ADDRESS, email_credentials.PASSWORD)
            login= 1
        except:
            print("Can not connect to the server.")
        if login == 1:
            server = self.Send_mail_body(server, month)
            server.quit()

    def Send_mail_body(self, server, month):
        names = list(self.user_data.Full_name)
        emails = list(self.user_data.Try_email)
        df_name = list(self.user_data.Df_name)
        number = 0
        for name, short_name, email in zip(names, df_name, emails):
            print(number, " :- " , name)
            msg = MIMEMultipart()

            msg['From'] = email_credentials.EMAIL_ADDRESS
            msg['To'] = email
            msg['Subject'] = "Result for the month of " + str(month)
            # add in the message body
            message = self.read_template("photos_and_other_requirement/format.txt").substitute(PERSON_NAME=name,MONTH_NAME=month)
            msg.attach(MIMEText(message, 'plain'))

            # open the file to be sent
            filename = short_name + "_" + str(month) + ".pdf"
            attachment = open("Report_card/" + str(month) + "/" + str(short_name) + "_" + str(month) + "/" + filename,
                              "rb")

            # instance of MIMEBase and named as p
            p = MIMEBase('application', 'octet-stream')

            # To change the payload into encoded form
            p.set_payload(attachment.read())

            # encode into base64
            encoders.encode_base64(p)

            p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

            # attach the instance 'p' to instance 'msg'
            msg.attach(p)

            server.send_message(msg)
            del (msg)

        return server

    def read_template(self, filename):
        with open(filename, 'r', encoding='utf-8') as template_file:
            template_file_content = template_file.read()
        return Template(template_file_content)


Main_call = Main()
